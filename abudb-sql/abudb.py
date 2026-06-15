import csv as _csv_mod
from datetime import date, datetime, timedelta


# ── Schema type coercion ──────────────────────────────────────────────────────

_TYPE_ALIASES = {
    'integer': 'int', 'number': 'float', 'text': 'string',
    'str': 'string', 'boolean': 'bool', 'duration': 'deltat',
}

def _coerce_to_schema_type(value, type_str, field_name='?'):
    """Coerce value to the declared datac schema type. Raises ValueError on failure."""
    t = _TYPE_ALIASES.get(type_str.lower(), type_str.lower())
    if t == 'identifier':
        t = 'int'   # IDENTIFIER is an auto-increment int; coerce same way
    if t == 'int':
        if isinstance(value, bool):
            raise ValueError(f"'{field_name}': bool is not int")
        if isinstance(value, int):
            return value
        if isinstance(value, float) and value == int(value):
            return int(value)
        if isinstance(value, str):
            try: return int(value)
            except ValueError: pass
        raise ValueError(f"'{field_name}': cannot coerce {value!r} to int")
    if t == 'float':
        if isinstance(value, bool):
            raise ValueError(f"'{field_name}': bool is not float")
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            try: return float(value)
            except ValueError: pass
        raise ValueError(f"'{field_name}': cannot coerce {value!r} to float")
    if t == 'string':
        return str(value)
    if t == 'bool':
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            if value.lower() == 'true':  return True
            if value.lower() == 'false': return False
        raise ValueError(f"'{field_name}': cannot coerce {value!r} to bool")
    if t == 'date':
        if isinstance(value, datetime):
            raise ValueError(f"'{field_name}': datetime is not date")
        if isinstance(value, date):
            return value
        raise ValueError(f"'{field_name}': expected date, got {type(value).__name__!r}")
    if t == 'datetime':
        if isinstance(value, datetime):
            return value
        raise ValueError(f"'{field_name}': expected datetime, got {type(value).__name__!r}")
    if t == 'deltat':
        if isinstance(value, timedelta):
            return value
        raise ValueError(f"'{field_name}': expected deltat, got {type(value).__name__!r}")
    return value   # unknown type — pass through


# ── Type coercion for CSV / XLSX (strings → typed) ───────────────────────────

def _coerce(v):
    if v is None or v == '':
        return None
    if isinstance(v, (int, float, bool)):
        return v
    s = str(v).strip()
    if not s:
        return None
    try:    return int(s)
    except (ValueError, TypeError): pass
    try:    return float(s)
    except (ValueError, TypeError): pass
    return s


# ── Source streamers (module-level, used by Table._stream) ───────────────────

def _stream_csv(path):
    with open(path, newline='', encoding='utf-8-sig') as f:
        reader = _csv_mod.DictReader(f)
        for row in reader:
            yield {k: _coerce(v) for k, v in row.items()}


def _stream_xlsx(path):
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl not installed — pip install openpyxl")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = None
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [str(h) if h is not None else f'col{i}'
                       for i, h in enumerate(row)]
        else:
            yield {h: _coerce(v) for h, v in zip(headers, row)}
    wb.close()


# ── Table ─────────────────────────────────────────────────────────────────────

class Table:
    def __init__(self, name, schema=None, primary_key=None, nullable=None, unique=None):
        self.name        = name
        self.schema      = schema or {}
        self.primary_key = primary_key
        self.nullable    = nullable or set()
        self.unique      = unique  or set()   # fields with ! unique constraint
        self._rows       = []      # in-memory rows; None = not yet loaded
        self._source     = None    # (path, kind[, table_name]) for lazy tables

    # ── Row access ───────────────────────────────────────────────────────────

    @property
    def rows(self):
        """Full in-memory list. Triggers complete load if table is lazy."""
        if self._source is not None:
            self._rows   = list(self._stream())
            self._source = None   # now fully eager
        return self._rows

    @rows.setter
    def rows(self, v):
        self._rows   = v
        self._source = None       # explicit assignment → fully eager

    @property
    def is_lazy(self):
        return self._source is not None

    def iter_rows(self, predicate=None):
        """
        Stream rows with optional filter applied at the source.
        For lazy tables this avoids loading non-matching rows into memory.
        """
        src = self._stream() if self._source is not None else self._rows
        for r in src:
            if predicate is None or predicate(r):
                yield r

    def _stream(self):
        path, kind = self._source[0], self._source[1]
        extra = self._source[2] if len(self._source) > 2 else None
        if kind == 'datac':
            from datac_parser import stream_datac_rows
            yield from stream_datac_rows(path, extra or self.name)
        elif kind == 'csv':
            yield from _stream_csv(path)
        elif kind == 'xlsx':
            yield from _stream_xlsx(path)

    # ── Copy ─────────────────────────────────────────────────────────────────

    def next_id(self, field):
        """Return the next auto-increment integer for an IDENTIFIER field."""
        existing = [r[field] for r in self._rows if isinstance(r.get(field), int)]
        return max(existing, default=0) + 1

    def copy(self):
        t = Table(self.name, self.schema.copy(), self.primary_key,
                  set(self.nullable), set(self.unique))
        if self._source is not None:
            t._source = self._source   # lazy copy — same source, no load
        else:
            t._rows = [r.copy() for r in self._rows]
        return t

    # ── Schema validation ─────────────────────────────────────────────────────

    def coerce_field(self, field, value):
        """Coerce a single value to the schema type for field. Raises ValueError."""
        if not self.schema or field not in self.schema:
            return value
        if value is None:
            if field not in self.nullable:
                raise ValueError(f"Field '{field}' is not nullable")
            return None
        return _coerce_to_schema_type(value, self.schema[field], field)

    def validate_and_coerce(self, row):
        """Coerce all schema fields in row. Returns new row dict. Raises ValueError."""
        if not self.schema:
            return row
        out = {}
        for field, type_str in self.schema.items():
            if field in row:
                val = row[field]
                if val is None:
                    if field not in self.nullable:
                        raise ValueError(f"Field '{field}' is not nullable")
                    out[field] = None
                else:
                    out[field] = _coerce_to_schema_type(val, type_str, field)
            elif field in self.nullable:
                out[field] = None
        for field, val in row.items():
            if field not in self.schema:
                out[field] = val
        return out

    # ── ALTER TABLE ───────────────────────────────────────────────────────────

    def alter_add_field(self, field, type_str, nullable=False, default=None):
        if field in self.schema:
            raise KeyError(f"Field '{field}' already exists in '{self.name}'")
        self.schema[field] = type_str
        if nullable:
            self.nullable.add(field)
        for row in self.rows:
            if field not in row:
                row[field] = default

    def alter_drop_field(self, field):
        if field not in self.schema:
            raise KeyError(f"Field '{field}' not found in '{self.name}'")
        if field == self.primary_key:
            raise ValueError(f"Cannot drop primary key field '{field}'")
        del self.schema[field]
        self.nullable.discard(field)
        for row in self.rows:
            row.pop(field, None)

    def alter_rename_field(self, old_field, new_field):
        if old_field not in self.schema:
            raise KeyError(f"Field '{old_field}' not found in '{self.name}'")
        if new_field in self.schema:
            raise KeyError(f"Field '{new_field}' already exists in '{self.name}'")
        # Preserve insertion order
        self.schema = {(new_field if k == old_field else k): v
                       for k, v in self.schema.items()}
        if old_field in self.nullable:
            self.nullable.discard(old_field)
            self.nullable.add(new_field)
        if self.primary_key == old_field:
            self.primary_key = new_field
        for row in self.rows:
            if old_field in row:
                row[new_field] = row.pop(old_field)


# ── Database ──────────────────────────────────────────────────────────────────

class Database:
    def __init__(self):
        self.tables          = {}
        self.indexes         = {}    # (table_name, field) → {value: [row_indices]}
        self.originals       = {}    # table_name → Table snapshot before CONNECT
        self.connections     = {}    # (left_name, right_name) → (left_field, right_field)
        self.views           = {}    # name → TAKE query string
        self.fk_constraints  = {}    # (child_table, child_field) → (parent_table, parent_field)
        self.checks          = {}    # table_name → [condition_strings]
        self._in_transaction = False
        self._tx_snapshot    = {}
        self._checkpoints    = {}    # name → snapshot dict (within transaction)

    # ── Import ────────────────────────────────────────────────────────────────

    def import_datac(self, filepath, alias=None):
        """Lazy import: reads schema headers eagerly, streams rows on demand."""
        from datac_parser import parse_datac_schemas
        schemas  = parse_datac_schemas(filepath)
        imported = []
        for schema, pk, name, nullable in schemas:
            tname = alias if (alias and len(schemas) == 1) else name
            t = Table(tname, schema, pk, nullable)
            t._source = (filepath, 'datac', name)   # stream from file when needed
            self.tables[tname] = t
            imported.append(t)
        return imported

    def import_csv(self, filepath, alias=None):
        """Lazy CSV import: reads header row eagerly, streams data on demand."""
        with open(filepath, newline='', encoding='utf-8-sig') as f:
            reader  = _csv_mod.DictReader(f)
            headers = reader.fieldnames or []
        tname  = alias or filepath.rsplit('/', 1)[-1].rsplit('.', 1)[0]
        schema = {h: 'string' for h in headers}
        t = Table(tname, schema)
        t._source = (filepath, 'csv')
        self.tables[tname] = t
        return t

    def import_xlsx(self, filepath, alias=None):
        """Lazy XLSX import: reads header row eagerly, streams data on demand."""
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError("openpyxl not installed — pip install openpyxl")
        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb.active
        headers = [str(c.value) if c.value is not None else f'col{i}'
                   for i, c in enumerate(next(ws.iter_rows()))]
        wb.close()
        tname  = alias or filepath.rsplit('/', 1)[-1].rsplit('.', 1)[0]
        schema = {h: 'string' for h in headers}
        t = Table(tname, schema)
        t._source = (filepath, 'xlsx')
        self.tables[tname] = t
        return t

    def content_from(self, filepath, alias):
        """Eager load into memory (legacy; IMPORT preferred for large files)."""
        if filepath.endswith('.csv'):
            t = self.import_csv(filepath, alias)
            t.rows  # trigger full load
        elif filepath.endswith('.xlsx'):
            t = self.import_xlsx(filepath, alias)
            t.rows
        else:
            raise ValueError(f"Unsupported file type: {filepath}")

    # ── DDL ───────────────────────────────────────────────────────────────────

    def create_table(self, name, schema, primary_key=None, nullable=None, unique=None):
        if name in self.tables:
            raise KeyError(f"Table '{name}' already exists")
        t = Table(name, schema, primary_key, nullable or set(), unique or set())
        t._rows = []
        self.tables[name] = t
        return t

    def drop_table(self, name):
        if name not in self.tables:
            raise KeyError(f"Table '{name}' not found")
        del self.tables[name]
        for key in list(self.indexes):
            if key[0] == name:
                del self.indexes[key]

    # ── Connect ───────────────────────────────────────────────────────────────

    # ── Views ─────────────────────────────────────────────────────────────────

    def create_view(self, name, query):
        if name in self.tables:
            raise KeyError(f"Cannot create view '{name}': a table with that name exists")
        self.views[name] = query

    def drop_view(self, name):
        if name not in self.views:
            raise KeyError(f"View '{name}' not found")
        del self.views[name]

    # ── Foreign keys ──────────────────────────────────────────────────────────

    def check_fk_insert(self, table_name, row):
        """Raise ValueError if any FK constraint is violated for a new row."""
        for (ct, cf), (pt, pf) in self.fk_constraints.items():
            if ct == table_name and cf in row:
                val = row[cf]
                if val is not None:
                    parent = self.tables.get(pt)
                    if parent and not any(r.get(pf) == val for r in parent.iter_rows()):
                        raise ValueError(
                            f"FK violation: {ct}.{cf}={val!r} references {pt}.{pf} which does not exist"
                        )

    def check_fk_delete(self, table_name, rows_being_deleted):
        """Raise ValueError if deleting rows would leave orphaned child rows."""
        deleting_vals = {r.get(f) for (pt, f), _ in [] for r in rows_being_deleted}
        for (ct, cf), (pt, pf) in self.fk_constraints.items():
            if pt == table_name:
                del_vals = {r.get(pf) for r in rows_being_deleted if r.get(pf) is not None}
                if not del_vals:
                    continue
                child = self.tables.get(ct)
                if child:
                    orphaned = any(r.get(cf) in del_vals for r in child.iter_rows())
                    if orphaned:
                        raise ValueError(
                            f"FK violation: deleting from '{table_name}' would orphan rows in '{ct}.{cf}'"
                        )

    # ── Connect ───────────────────────────────────────────────────────────────

    def connect(self, table1_name, field1, table2_name, field2):
        t1 = self.tables.get(table1_name)
        t2 = self.tables.get(table2_name)
        if not t1: raise KeyError(f"Table '{table1_name}' not found")
        if not t2: raise KeyError(f"Table '{table2_name}' not found")
        # table2.field2 references table1.field1
        self.fk_constraints[(table2_name, field2)] = (table1_name, field1)

        if table1_name not in self.originals:
            self.originals[table1_name] = t1.copy()
        if table2_name not in self.originals:
            self.originals[table2_name] = t2.copy()

        self.connections[(table1_name, table2_name)] = (field1, field2)
        self.connections[(table2_name, table1_name)] = (field2, field1)

        t2_lookup = {}
        for row in t2.rows:
            key = row.get(field2)
            if key is not None:
                t2_lookup.setdefault(key, []).append(row)

        conflicts   = (set(t1.schema) & set(t2.schema)) - {field1, field2}
        merged_rows = []
        for r1 in t1.rows:
            for r2 in t2_lookup.get(r1.get(field1), []):
                merged = {}
                for k, v in r1.items():
                    merged[k] = v
                    merged[f'{table1_name}.{k}'] = v
                for k, v in r2.items():
                    if k in conflicts:
                        merged[f'{table2_name}.{k}'] = v
                    else:
                        merged[k] = v
                    merged[f'{table2_name}.{k}'] = v
                merged_rows.append(merged)

        merged_schema = {**t1.schema, **t2.schema}
        mt = Table(f'{table1_name}+{table2_name}', merged_schema)
        mt.rows = merged_rows
        self.tables[table1_name] = mt
        self.tables[table2_name] = mt
        self.tables[mt.name]     = mt
        return mt

    # ── Left join (KEEP ALL) ──────────────────────────────────────────────────

    def left_join(self, left_name, right_name):
        left_t  = self.originals.get(left_name)  or self.tables.get(left_name)
        right_t = self.originals.get(right_name) or self.tables.get(right_name)
        if left_t  is None: raise KeyError(f"Table '{left_name}' not found")
        if right_t is None: raise KeyError(f"Table '{right_name}' not found")

        conn = self.connections.get((left_name, right_name))
        if conn is None:
            raise KeyError(
                f"No CONNECT relationship between '{left_name}' and '{right_name}'. "
                "Run CONNECT first."
            )
        left_field, right_field = conn

        right_lookup = {}
        for row in right_t.rows:
            key = row.get(right_field)
            right_lookup.setdefault(key, []).append(row)

        right_fields = (list(right_t.schema.keys()) if right_t.schema else
                        list(right_t.rows[0].keys()) if right_t.rows else [])

        result = []
        for r1 in left_t.rows:
            matches = right_lookup.get(r1.get(left_field), [])
            if matches:
                for r2 in matches:
                    merged = dict(r1)
                    for k, v in r1.items():
                        merged[f'{left_name}.{k}'] = v
                    for k, v in r2.items():
                        merged[f'{right_name}.{k}'] = v
                        if k not in merged:
                            merged[k] = v
                    result.append(merged)
            else:
                merged = dict(r1)
                for k, v in r1.items():
                    merged[f'{left_name}.{k}'] = v
                for f in right_fields:
                    merged[f] = None
                    merged[f'{right_name}.{f}'] = None
                result.append(merged)
        return result

    # ── Indexes ───────────────────────────────────────────────────────────────

    def build_index(self, table_name, field):
        if table_name not in self.tables:
            raise KeyError(f"Table '{table_name}' not found")
        idx = {}
        for i, row in enumerate(self.tables[table_name].rows):
            key = row.get(field)
            idx.setdefault(key, []).append(i)
        self.indexes[(table_name, field)] = idx
        return len(idx)

    def drop_index(self, table_name, field):
        return self.indexes.pop((table_name, field), None) is not None

    def get_indexed_rows(self, table_name, field, value):
        key = (table_name, field)
        if key not in self.indexes:
            return None
        table   = self.tables[table_name]
        indices = self.indexes[key].get(value, [])
        return [table.rows[i] for i in indices if i < len(table.rows)]

    def invalidate_indexes(self, table_name):
        for (tn, field) in list(self.indexes):
            if tn == table_name:
                self.build_index(tn, field)

    # ── Transactions ──────────────────────────────────────────────────────────

    def begin(self):
        if self._in_transaction:
            raise RuntimeError("Already in a transaction — COMMIT or ROLLBACK first")
        self._tx_snapshot = self._make_snapshot()
        self._checkpoints    = {}
        self._in_transaction = True

    def commit(self):
        if not self._in_transaction:
            raise RuntimeError("No active transaction")
        self._tx_snapshot    = {}
        self._checkpoints    = {}
        self._in_transaction = False

    def rollback(self):
        snap = self._tx_snapshot
        if snap:
            self._restore_snapshot(snap)
        self._tx_snapshot    = {}
        self._checkpoints    = {}
        self._in_transaction = False

    def checkpoint(self, name):
        if not self._in_transaction:
            raise RuntimeError("CHECKPOINT requires an active transaction (BEGIN first)")
        self._checkpoints[name] = self._make_snapshot()

    def rollback_to(self, name):
        if name not in self._checkpoints:
            raise KeyError(f"Checkpoint '{name}' not found")
        self._restore_snapshot(self._checkpoints[name])
        # Drop checkpoints created after this one (not tracked by order — clear all after)
        keys = list(self._checkpoints)
        drop = False
        for k in keys:
            if drop:
                del self._checkpoints[k]
            if k == name:
                drop = True

    def _make_snapshot(self):
        return {
            'tables':         {n: t.copy() for n, t in self.tables.items()},
            'indexes':        dict(self.indexes),
            'originals':      {n: t.copy() for n, t in self.originals.items()},
            'connections':    dict(self.connections),
            'views':          dict(self.views),
            'fk_constraints': dict(self.fk_constraints),
            'checks':         {k: list(v) for k, v in self.checks.items()},
        }

    def _restore_snapshot(self, snap):
        self.tables         = snap['tables']
        self.indexes        = snap['indexes']
        self.originals      = snap['originals']
        self.connections    = snap['connections']
        self.views          = snap.get('views', {})
        self.fk_constraints = snap.get('fk_constraints', {})
        self.checks         = snap.get('checks', {})

    # ── ALTER TABLE (rename) ──────────────────────────────────────────────────

    def alter_rename_table(self, old_name, new_name):
        if old_name not in self.tables:
            raise KeyError(f"Table '{old_name}' not found")
        if new_name in self.tables:
            raise KeyError(f"Table '{new_name}' already exists")
        t = self.tables.pop(old_name)
        t.name = new_name
        self.tables[new_name] = t
        for (tn, field) in list(self.indexes):
            if tn == old_name:
                self.indexes[(new_name, field)] = self.indexes.pop((tn, field))
